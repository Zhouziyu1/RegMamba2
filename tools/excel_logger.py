# -*- coding:UTF-8 -*-
"""
Excel 训练日志记录器
================================================================================
设计者：周女士

功能：
1. 记录每个epoch的所有损失和指标
2. 横向：各项指标
3. 纵向：训练轮次
4. 自动保存为Excel文件，方便论文分析
================================================================================
"""

import os
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Any


class ExcelLogger:
    """
    Excel 训练日志记录器
    设计者：周女士
    
    使用方法:
        logger = ExcelLogger(save_dir='./experiment/logs', exp_name='RegMamba_KITTI')
        
        # 每个epoch记录
        logger.log_epoch(
            epoch=1,
            train_metrics={...},
            val_metrics={...},
        )
        
        # 保存
        logger.save()
    """
    
    def __init__(
        self,
        save_dir: str,
        exp_name: str = 'experiment',
        auto_save: bool = True,
    ):
        """
        Args:
            save_dir: 保存目录
            exp_name: 实验名称
            auto_save: 是否每次记录后自动保存
        """
        self.save_dir = save_dir
        self.exp_name = exp_name
        self.auto_save = auto_save
        
        os.makedirs(save_dir, exist_ok=True)
        
        # 时间戳
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Excel文件路径
        self.excel_path = os.path.join(save_dir, f'{exp_name}_{self.timestamp}.xlsx')
        
        # 数据存储
        self.train_data = []  # 训练数据
        self.val_data = []    # 验证数据
        self.test_data = []   # 测试数据
        
        # 定义所有需要记录的列（论文常用指标）
        self.train_columns = [
            # 基本信息
            'Epoch',
            'Learning_Rate',
            'Time_per_Epoch(s)',
            
            # ===== 损失函数 =====
            'Total_Loss',
            'Rotation_Loss',           # L_rot
            'Translation_Loss',        # L_trans
            'Overlap_Loss',            # L_overlap
            'DeepSupervision_Loss',    # L_ds (深度监督)
            'Feature_Loss_Layer1',     # 逐层特征损失
            'Feature_Loss_Layer2',
            'Feature_Loss_Layer3',
            'Feature_Loss_Layer4',
        ]
        
        self.val_columns = [
            # 基本信息
            'Epoch',
            'Val_Loss',
            
            # ===== 配准召回率 (Registration Recall) =====
            'RR@5°2m(%)',              # 默认阈值
            'RR@1°0.1m(%)',            # 严格阈值 (3DMatch常用)
            'RR@2°0.5m(%)',            # 中等阈值
            'RR@10°5m(%)',             # 宽松阈值
            
            # ===== 旋转误差 (Rotation Error) =====
            'RRE_Mean(°)',             # Relative Rotation Error 均值
            'RRE_Std(°)',              # 标准差
            'RRE_Median(°)',           # 中位数
            'RRE_Min(°)',              # 最小值
            'RRE_Max(°)',              # 最大值
            'RRE_RMSE(°)',             # 均方根误差
            
            # ===== 平移误差 (Translation Error) =====
            'RTE_Mean(m)',             # Relative Translation Error 均值
            'RTE_Std(m)',              # 标准差
            'RTE_Median(m)',           # 中位数
            'RTE_Min(m)',              # 最小值
            'RTE_Max(m)',              # 最大值
            'RTE_RMSE(m)',             # 均方根误差
            
            # ===== 成功样本统计 =====
            'Num_Success',             # 成功样本数
            'Num_Total',               # 总样本数
            'Success_RRE_Mean(°)',     # 成功样本的RRE均值
            'Success_RTE_Mean(m)',     # 成功样本的RTE均值
            
            # ===== 点级误差 =====
            'Point_RMSE(m)',           # 点云RMSE
            'Point_MAE(m)',            # 点云MAE
            'Chamfer_Distance',        # 倒角距离
            
            # ===== 其他论文常用 =====
            'Inlier_Ratio(%)',         # 内点比例
            'Feature_Match_Recall(%)', # 特征匹配召回率
            
            # ===== 推理速度 =====
            'Inference_Time(ms)',      # 推理时间
            'FPS',                     # 帧率
        ]
        
        print(f"[ExcelLogger] 初始化完成")
        print(f"  - 保存路径: {self.excel_path}")
        print(f"  - 训练指标列数: {len(self.train_columns)}")
        print(f"  - 验证指标列数: {len(self.val_columns)}")
    
    def log_train_epoch(
        self,
        epoch: int,
        learning_rate: float,
        time_elapsed: float,
        total_loss: float,
        rot_loss: float,
        trans_loss: float,
        overlap_loss: float = 0.0,
        ds_loss: float = 0.0,
        feature_losses: List[float] = None,
        **kwargs,
    ):
        """
        记录训练epoch数据
        
        Args:
            epoch: 当前轮次
            learning_rate: 学习率
            time_elapsed: 本轮耗时(秒)
            total_loss: 总损失
            rot_loss: 旋转损失
            trans_loss: 平移损失
            overlap_loss: 重叠损失
            ds_loss: 深度监督损失
            feature_losses: 各层特征损失 [layer1, layer2, ...]
            **kwargs: 其他自定义指标
        """
        # 处理特征损失
        if feature_losses is None:
            feature_losses = [0.0, 0.0, 0.0, 0.0]
        while len(feature_losses) < 4:
            feature_losses.append(0.0)
        
        row = {
            'Epoch': epoch,
            'Learning_Rate': learning_rate,
            'Time_per_Epoch(s)': time_elapsed,
            'Total_Loss': total_loss,
            'Rotation_Loss': rot_loss,
            'Translation_Loss': trans_loss,
            'Overlap_Loss': overlap_loss,
            'DeepSupervision_Loss': ds_loss,
            'Feature_Loss_Layer1': feature_losses[0],
            'Feature_Loss_Layer2': feature_losses[1],
            'Feature_Loss_Layer3': feature_losses[2],
            'Feature_Loss_Layer4': feature_losses[3],
        }
        
        # 添加额外的自定义指标
        row.update(kwargs)
        
        self.train_data.append(row)
        
        if self.auto_save:
            self.save()
    
    def log_val_epoch(
        self,
        epoch: int,
        val_loss: float,
        rot_errors: np.ndarray,
        trans_errors: np.ndarray,
        rot_thresh: float = 5.0,
        trans_thresh: float = 2.0,
        point_rmse: float = None,
        point_mae: float = None,
        chamfer_dist: float = None,
        inlier_ratio: float = None,
        feature_match_recall: float = None,
        inference_time: float = None,
        **kwargs,
    ):
        """
        记录验证epoch数据
        
        Args:
            epoch: 当前轮次
            val_loss: 验证损失
            rot_errors: 所有样本的旋转误差数组 (N,)
            trans_errors: 所有样本的平移误差数组 (N,)
            rot_thresh: 默认旋转阈值
            trans_thresh: 默认平移阈值
            point_rmse: 点级RMSE
            point_mae: 点级MAE
            chamfer_dist: Chamfer距离
            inlier_ratio: 内点比例
            feature_match_recall: 特征匹配召回率
            inference_time: 推理时间(ms)
        """
        rot_errors = np.array(rot_errors)
        trans_errors = np.array(trans_errors)
        N = len(rot_errors)
        
        # ===== 计算各阈值下的召回率 =====
        def calc_recall(r_thresh, t_thresh):
            success = (rot_errors < r_thresh) & (trans_errors < t_thresh)
            return np.mean(success) * 100
        
        rr_5_2 = calc_recall(5.0, 2.0)      # 默认
        rr_1_01 = calc_recall(1.0, 0.1)     # 严格
        rr_2_05 = calc_recall(2.0, 0.5)     # 中等
        rr_10_5 = calc_recall(10.0, 5.0)    # 宽松
        
        # ===== 成功样本 =====
        success_mask = (rot_errors < rot_thresh) & (trans_errors < trans_thresh)
        num_success = np.sum(success_mask)
        success_indices = np.where(success_mask)[0]
        
        if num_success > 0:
            success_rre_mean = np.mean(rot_errors[success_indices])
            success_rte_mean = np.mean(trans_errors[success_indices])
        else:
            success_rre_mean = 0.0
            success_rte_mean = 0.0
        
        # ===== 构建记录行 =====
        row = {
            'Epoch': epoch,
            'Val_Loss': val_loss,
            
            # 召回率
            'RR@5°2m(%)': rr_5_2,
            'RR@1°0.1m(%)': rr_1_01,
            'RR@2°0.5m(%)': rr_2_05,
            'RR@10°5m(%)': rr_10_5,
            
            # 旋转误差
            'RRE_Mean(°)': np.mean(rot_errors),
            'RRE_Std(°)': np.std(rot_errors),
            'RRE_Median(°)': np.median(rot_errors),
            'RRE_Min(°)': np.min(rot_errors),
            'RRE_Max(°)': np.max(rot_errors),
            'RRE_RMSE(°)': np.sqrt(np.mean(rot_errors ** 2)),
            
            # 平移误差
            'RTE_Mean(m)': np.mean(trans_errors),
            'RTE_Std(m)': np.std(trans_errors),
            'RTE_Median(m)': np.median(trans_errors),
            'RTE_Min(m)': np.min(trans_errors),
            'RTE_Max(m)': np.max(trans_errors),
            'RTE_RMSE(m)': np.sqrt(np.mean(trans_errors ** 2)),
            
            # 成功样本统计
            'Num_Success': num_success,
            'Num_Total': N,
            'Success_RRE_Mean(°)': success_rre_mean,
            'Success_RTE_Mean(m)': success_rte_mean,
            
            # 点级误差
            'Point_RMSE(m)': point_rmse if point_rmse is not None else 0.0,
            'Point_MAE(m)': point_mae if point_mae is not None else 0.0,
            'Chamfer_Distance': chamfer_dist if chamfer_dist is not None else 0.0,
            
            # 其他
            'Inlier_Ratio(%)': inlier_ratio if inlier_ratio is not None else 0.0,
            'Feature_Match_Recall(%)': feature_match_recall if feature_match_recall is not None else 0.0,
            
            # 速度
            'Inference_Time(ms)': inference_time if inference_time is not None else 0.0,
            'FPS': 1000.0 / inference_time if inference_time and inference_time > 0 else 0.0,
        }
        
        # 添加额外指标
        row.update(kwargs)
        
        self.val_data.append(row)
        
        if self.auto_save:
            self.save()
    
    def log_test(
        self,
        test_name: str,
        rot_errors: np.ndarray,
        trans_errors: np.ndarray,
        **kwargs,
    ):
        """记录测试结果"""
        rot_errors = np.array(rot_errors)
        trans_errors = np.array(trans_errors)
        N = len(rot_errors)
        
        row = {
            'Test_Name': test_name,
            'Num_Samples': N,
            
            # 各阈值召回率
            'RR@5°2m(%)': np.mean((rot_errors < 5) & (trans_errors < 2)) * 100,
            'RR@1°0.1m(%)': np.mean((rot_errors < 1) & (trans_errors < 0.1)) * 100,
            'RR@2°0.5m(%)': np.mean((rot_errors < 2) & (trans_errors < 0.5)) * 100,
            
            # 误差统计
            'RRE_Mean(°)': np.mean(rot_errors),
            'RRE_Median(°)': np.median(rot_errors),
            'RTE_Mean(m)': np.mean(trans_errors),
            'RTE_Median(m)': np.median(trans_errors),
        }
        row.update(kwargs)
        
        self.test_data.append(row)
        
        if self.auto_save:
            self.save()
    
    def save(self):
        """保存到Excel文件"""
        with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
            # 训练数据
            if self.train_data:
                df_train = pd.DataFrame(self.train_data)
                df_train.to_excel(writer, sheet_name='Training', index=False)
                self._format_sheet(writer, 'Training', df_train)
            
            # 验证数据
            if self.val_data:
                df_val = pd.DataFrame(self.val_data)
                df_val.to_excel(writer, sheet_name='Validation', index=False)
                self._format_sheet(writer, 'Validation', df_val)
            
            # 测试数据
            if self.test_data:
                df_test = pd.DataFrame(self.test_data)
                df_test.to_excel(writer, sheet_name='Test', index=False)
                self._format_sheet(writer, 'Test', df_test)
            
            # 摘要sheet
            self._create_summary_sheet(writer)
        
        # print(f"[ExcelLogger] 已保存: {self.excel_path}")
    
    def _format_sheet(self, writer, sheet_name: str, df: pd.DataFrame):
        """格式化Excel工作表"""
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            ws = writer.sheets[sheet_name]
            
            # 表头样式
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font_white = Font(bold=True, size=11, color='FFFFFF')
            
            # 边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置表头
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # 设置数据单元格
            for row_num in range(2, len(df) + 2):
                for col_num in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    # 数值格式化
                    if isinstance(cell.value, float):
                        cell.number_format = '0.0000'
            
            # 自动调整列宽
            for col_num, column_title in enumerate(df.columns, 1):
                max_length = max(
                    len(str(column_title)),
                    df[column_title].astype(str).str.len().max() if len(df) > 0 else 0
                )
                ws.column_dimensions[chr(64 + col_num) if col_num <= 26 else f'{chr(64 + col_num // 26)}{chr(64 + col_num % 26)}'].width = min(max_length + 2, 20)
                
        except ImportError:
            pass  # openpyxl 未安装样式模块
    
    def _create_summary_sheet(self, writer):
        """创建摘要sheet"""
        if not self.val_data:
            return
        
        df_val = pd.DataFrame(self.val_data)
        
        # 找最优epoch
        best_recall_idx = df_val['RR@5°2m(%)'].idxmax()
        best_loss_idx = df_val['Val_Loss'].idxmin()
        
        summary_data = {
            '指标': [
                '实验名称',
                '总训练轮数',
                '最优验证Loss',
                '最优Loss对应Epoch',
                '最高召回率 RR@5°2m',
                '最高召回率对应Epoch',
                '最终RRE均值',
                '最终RTE均值',
            ],
            '数值': [
                self.exp_name,
                len(self.train_data) if self.train_data else 0,
                df_val.loc[best_loss_idx, 'Val_Loss'],
                df_val.loc[best_loss_idx, 'Epoch'],
                df_val.loc[best_recall_idx, 'RR@5°2m(%)'],
                df_val.loc[best_recall_idx, 'Epoch'],
                df_val.iloc[-1]['RRE_Mean(°)'],
                df_val.iloc[-1]['RTE_Mean(m)'],
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def get_best_metrics(self) -> Dict[str, Any]:
        """获取最优指标"""
        if not self.val_data:
            return {}
        
        df_val = pd.DataFrame(self.val_data)
        best_idx = df_val['RR@5°2m(%)'].idxmax()
        
        return {
            'best_epoch': df_val.loc[best_idx, 'Epoch'],
            'best_recall': df_val.loc[best_idx, 'RR@5°2m(%)'],
            'best_rre': df_val.loc[best_idx, 'RRE_Mean(°)'],
            'best_rte': df_val.loc[best_idx, 'RTE_Mean(m)'],
        }


# =============================================================================
# 测试
# =============================================================================

if __name__ == "__main__":
    print("测试 ExcelLogger...")
    
    np.random.seed(42)
    
    # 创建logger
    logger = ExcelLogger(
        save_dir='./test_logs',
        exp_name='RegMamba_KITTI_test',
        auto_save=True,
    )
    
    # 模拟10个epoch
    for epoch in range(1, 11):
        # 训练数据
        logger.log_train_epoch(
            epoch=epoch,
            learning_rate=0.001 * (0.9 ** epoch),
            time_elapsed=120.5 + np.random.randn() * 10,
            total_loss=1.0 / epoch + np.random.randn() * 0.1,
            rot_loss=0.5 / epoch + np.random.randn() * 0.05,
            trans_loss=0.3 / epoch + np.random.randn() * 0.03,
            overlap_loss=0.1 / epoch,
            ds_loss=0.1 / epoch,
            feature_losses=[0.1/epoch, 0.08/epoch, 0.06/epoch, 0.04/epoch],
        )
        
        # 验证数据 (每2个epoch)
        if epoch % 2 == 0:
            rot_errors = np.abs(np.random.randn(100) * (5.0 / epoch))
            trans_errors = np.abs(np.random.randn(100) * (2.0 / epoch))
            
            logger.log_val_epoch(
                epoch=epoch,
                val_loss=0.8 / epoch,
                rot_errors=rot_errors,
                trans_errors=trans_errors,
                point_rmse=0.1 / epoch,
                point_mae=0.08 / epoch,
                inference_time=50.0,
            )
    
    print(f"\n保存路径: {logger.excel_path}")
    print(f"最优指标: {logger.get_best_metrics()}")
    print("\n测试完成！✓")